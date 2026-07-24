"""
Exporter utility — generates Excel workbook from filtered promotions DataFrame.
"""
import io
from collections.abc import Mapping, Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


WEEKDAY_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
UNASSIGNED_TEAM_ID = "unassigned"
UNASSIGNED_TEAM_NAME = "Unassigned / General"


def _team_ids_from_value(team_ids_value) -> list[str]:
    """Return normalized team IDs from the comma-separated DB aggregate."""
    if team_ids_value is None or pd.isna(team_ids_value) or not str(team_ids_value).strip():
        return []
    return [team_id.strip() for team_id in str(team_ids_value).split(",") if team_id.strip()]


def _team_dataframe(df: pd.DataFrame, team_id: str) -> pd.DataFrame:
    if team_id == UNASSIGNED_TEAM_ID:
        return df[df["team_ids_str"].apply(lambda value: not _team_ids_from_value(value))].copy()

    return df[df["team_ids_str"].apply(lambda value: team_id in _team_ids_from_value(value))].copy()


def _dedupe_offer_lines(values: pd.Series) -> str:
    return "\n".join(dict.fromkeys(v for v in values if isinstance(v, str) and v.strip()))


def _default_team_ids(df: pd.DataFrame) -> list[str]:
    team_ids: list[str] = []
    has_unassigned = False

    for value in df.get("team_ids_str", pd.Series(dtype="str")):
        row_team_ids = _team_ids_from_value(value)
        if not row_team_ids:
            has_unassigned = True
            continue

        for team_id in row_team_ids:
            if team_id not in team_ids:
                team_ids.append(team_id)

    if has_unassigned:
        team_ids.append(UNASSIGNED_TEAM_ID)

    return team_ids


def export_to_excel(
    df: pd.DataFrame,
    selected_team_ids: Sequence[str] | None = None,
    team_map: Mapping[str, str] | None = None,
    allowed_brands_by_team: Mapping[str, list[str]] | None = None,
) -> bytes:
    """
    Exports the filtered promotions DataFrame to an Excel spreadsheet in the
    exact Weekly Competitor Matrix format shown on the Streamlit dashboard.

    Layout:
    - Grouped by selected business team, in the same order as the dashboard.
    - Team name acts as a section header row.
    - Grid columns: Team Brand (first column), Monday, Tuesday, Wednesday, Thursday, Friday, Saturday, Sunday.
    - Cell contents contain newline-separated promotion titles.
    - Soft styling matching the dashboard's design tokens.
    """
    if df.empty:
        wb = Workbook()
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    matrix_df = df.copy()
    if "team_ids_str" not in matrix_df:
        matrix_df["team_ids_str"] = ""

    selected_team_ids = list(selected_team_ids) if selected_team_ids is not None else _default_team_ids(matrix_df)
    team_map = dict(team_map or {})
    team_map.setdefault(UNASSIGNED_TEAM_ID, UNASSIGNED_TEAM_NAME)

    matrix_df["Day"] = matrix_df["scraped_at"].dt.day_name()
    _allowed = dict(allowed_brands_by_team or {})

    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Competitor Matrix"

    # Ensure grid lines are explicitly visible
    ws.views.sheetView[0].showGridLines = True

    # Define design colors and styles matching styles.py
    # Team Header Banner Style
    team_fill = PatternFill(start_color="EEF0F8", end_color="EEF0F8", fill_type="solid")
    team_font = Font(name="Segoe UI", size=11, bold=True, color="303247")

    # Table Column Header Style
    header_fill = PatternFill(start_color="F4F4F8", end_color="F4F4F8", fill_type="solid")
    header_font = Font(name="Segoe UI", size=10, bold=True, color="505060")

    # Row Headers (Brand Names) and Data Cells
    brand_font = Font(name="Segoe UI", size=10, bold=True, color="1A1A2E")
    data_font = Font(name="Segoe UI", size=10, bold=False, color="1A1A2E")

    # Soft borders
    border_side = Side(border_style="thin", color="E5E5EE")
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    # Alignments
    align_left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_center = Alignment(horizontal="center", vertical="center")

    current_row = 1

    # Group and build matrices for each selected business team.
    for team_id in selected_team_ids:
        team_name = team_map.get(team_id, team_id)
        team_df = _team_dataframe(matrix_df, team_id)
        configured_brands = _allowed.get(team_id, [])  # brands from teams.json

        brand_column = f"{team_name} Brand"

        if team_df.empty:
            # All configured brands with zero offers
            empty_rows = [{brand_column: b, **{d: "" for d in WEEKDAY_ORDER}} for b in configured_brands]
            matrix = pd.DataFrame(empty_rows) if empty_rows else pd.DataFrame(columns=[brand_column] + WEEKDAY_ORDER)
        else:
            # Group duplicates and join with newlines
            grouped = (
                team_df.groupby(["brand", "Day"])["offer_title"]
                .apply(_dedupe_offer_lines)
                .reset_index()
            )
            matrix = grouped.pivot(index="brand", columns="Day", values="offer_title")
            matrix = matrix.reindex(columns=WEEKDAY_ORDER).fillna("")
            matrix = matrix.reset_index().rename(columns={"brand": brand_column})

            # Inject configured brands that had zero offers in this period
            scraped_brands = set(matrix[brand_column].tolist())
            missing = [b for b in configured_brands if b not in scraped_brands]
            if missing:
                missing_rows = pd.DataFrame(
                    [{brand_column: b, **{d: "" for d in WEEKDAY_ORDER}} for b in missing]
                )
                matrix = pd.concat([matrix, missing_rows], ignore_index=True)

        # Sort alphabetically for consistent display
        matrix = matrix.sort_values(brand_column).reset_index(drop=True)

        if matrix.empty:
            continue

        n_brands = matrix.shape[0]
        n_promos = team_df["offer_title"].nunique() if not team_df.empty else 0

        # 1. Write Team Title Block (Merged Row)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        team_cell = ws.cell(row=current_row, column=1)
        team_cell.value = f" {team_name} | {n_brands} brand{'s' if n_brands != 1 else ''} | {n_promos} unique offer{'s' if n_promos != 1 else ''}"
        team_cell.font = team_font
        team_cell.fill = team_fill
        team_cell.alignment = Alignment(horizontal="left", vertical="center")

        for c in range(1, 9):
            ws.cell(row=current_row, column=c).border = cell_border
            ws.cell(row=current_row, column=c).fill = team_fill

        ws.row_dimensions[current_row].height = 28
        current_row += 1

        # 2. Write Weekday Column Headers
        headers = [brand_column] + WEEKDAY_ORDER
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center if col_idx > 1 else align_left_wrap
            cell.border = cell_border

        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # 3. Write Matrix Data Rows
        for _, row in matrix.iterrows():
            # Brand Name cell
            brand_cell = ws.cell(row=current_row, column=1)
            brand_cell.value = row[brand_column]
            brand_cell.font = brand_font
            brand_cell.alignment = align_left_wrap
            brand_cell.border = cell_border

            # Days of the Week cells
            max_lines = 1
            for d_idx, day in enumerate(WEEKDAY_ORDER, start=2):
                cell = ws.cell(row=current_row, column=d_idx)
                cell_value = row[day] if str(row[day]).strip() else "—"
                cell.value = cell_value
                cell.font = data_font
                cell.alignment = align_left_wrap
                cell.border = cell_border

                # Check line counts to calculate row height
                lines = len(cell_value.split('\n'))
                if lines > max_lines:
                    max_lines = lines

            # Set dynamic row height to fit wrapped text
            ws.row_dimensions[current_row].height = max(max_lines * 15 + 10, 30)
            current_row += 1

        # Spacing row after matrix table
        current_row += 1
        ws.row_dimensions[current_row].height = 15
        current_row += 1

    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 22
    for c in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[c].width = 30

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
