import io

import pandas as pd
from openpyxl import load_workbook

from dashboard.utils.exporter import export_to_excel


def test_export_to_excel_matches_team_matrix_view():
    df = pd.DataFrame(
        [
            {
                "brand": "Brand A",
                "offer_title": "20% off",
                "category": "Beauty",
                "scraped_at": pd.Timestamp("2026-07-20 09:00:00"),
                "team_ids_str": "beauty",
            },
            {
                "brand": "Brand A",
                "offer_title": "Free gift",
                "category": "Beauty",
                "scraped_at": pd.Timestamp("2026-07-20 10:00:00"),
                "team_ids_str": "beauty",
            },
            {
                "brand": "Brand B",
                "offer_title": "Half price",
                "category": "Home",
                "scraped_at": pd.Timestamp("2026-07-21 09:00:00"),
                "team_ids_str": "home",
            },
            {
                "brand": "Brand C",
                "offer_title": "Clearance",
                "category": "General",
                "scraped_at": pd.Timestamp("2026-07-22 09:00:00"),
                "team_ids_str": "",
            },
        ]
    )

    workbook_bytes = export_to_excel(
        df,
        selected_team_ids=["beauty", "unassigned"],
        team_map={"beauty": "Beauty", "home": "Home", "unassigned": "Unassigned / General"},
    )

    workbook = load_workbook(io.BytesIO(workbook_bytes))
    sheet = workbook["Weekly Competitor Matrix"]

    assert sheet["A1"].value == " Beauty | 1 brand | 2 unique offers"
    assert [sheet.cell(2, col).value for col in range(1, 9)] == [
        "Beauty Brand",
        "Monday",
        "Tuesday",
        "Wednesday",
        "Thursday",
        "Friday",
        "Saturday",
        "Sunday",
    ]
    assert sheet["A3"].value == "Brand A"
    assert sheet["B3"].value == "20% off\nFree gift"
    assert sheet["C3"].value == "—"

    assert sheet["A6"].value == " Unassigned / General | 1 brand | 1 unique offer"
    assert sheet["A8"].value == "Brand C"
    assert sheet["D8"].value == "Clearance"
    assert "Home" not in [sheet.cell(row, 1).value for row in range(1, sheet.max_row + 1)]
